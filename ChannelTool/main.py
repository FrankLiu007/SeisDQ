import sys
import os
import datetime
import time
import json
import csv
import openpyxl

from PyQt5 import QtWidgets, QtGui
from PyQt5.QtCore import Qt
from PyQt5 import QtCore
from PyQt5.QtWidgets import QMainWindow, QApplication, QTreeWidgetItem, QMessageBox,QProgressDialog,QProgressBar

QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True) #enable highdpi scaling
QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True) #use highdpi icons

from ChannelTool.window import Ui_MainWindow
from ChannelTool.station_spyder import crawl

# proxies = {
#     "http": "http://127.0.0.1:1082",
#     "https": "http://127.0.0.1:1082"
# }
proxies = None


class UiMain(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None):
        super(UiMain, self).__init__(parent)
        self.setupUi(self)
        self.component_list = {}
        self.currStationIndex=-1
        self.stations=[]
        self.currStation={}
        self.progressDialog=None
        self.hasData=False
        self.crawl_btn.setEnabled(False)
        self.crawl_btn.setStyleSheet('font: 10pt "Microsoft YaHei UI";background-color:#455ab3;color:#fff;')

        self.bind_signal()


    def bind_signal(self):
        self.read_excel_btn.clicked.connect(self.read_excel_btn_clicked)
        self.crawl_btn.clicked.connect(self.crawl_btn_clicked)
        self.next_btn.clicked.connect(self.next_btn_clicked)
        self.prior_btn.clicked.connect(self.prior_btn_clicked)

        self.stationListWidget.itemClicked.connect(self.station_item_clicked)
        self.treeWidget.itemChanged.connect(self.channelSelectionChanged)
        self.save_btn.clicked.connect( self.save_btn_clicked )
        self.export_cmp_btn.clicked.connect(self.export_cmp_btn_clicked)
        self.load_btn.clicked.connect(self.load_btn_clicked)

    def export_cmp_btn_clicked(self):
        fname, filetype = QtWidgets.QFileDialog.getSaveFileName(self, "导出台站信息", "", "csv Files (*.csv;)")
        if fname == "":
            return
        self.exportSelectedChannels(fname)

    def updateStationLabel(self,station):
        if "selectedChannels" in station:
            n=len(station["selectedChannels"])
        else:
            n=0
        self.stationListWidget.currentItem().setText(self.currStation["name"]+"/"+self.currStation["name"]+ " ( "+str(n)+"  channels selected" +" )")

    def channelSelectionChanged(self,item, col):
        if item.text(1):  ##item.text(1) is the key to channel object
            self.currStation["allChannels"][item.text(1)]["selected"]=item.checkState(0)
            self.updateSelectedChannels(item)
            self.updateStationLabel(self.currStation)

    def updateSelectedChannels(self, item):
        if item.checkState(0) ==Qt.Unchecked:
            self.currStation["selectedChannels"].pop(item.text(1))
        else:
            self.currStation["selectedChannels"][item.text(1)]=self.currStation["allChannels"][item.text(1)]

    def read_excel_btn_clicked(self):

        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(self, "选取文件", os.getcwd(), "Execl Files (*.xlsx;*.xls)")  # 设置文件扩展名过滤,用双分号间隔

        if filename == "":
            return
        # 重置软件状态
        self.reset_data()
        self.excel_path_edit.setText(filename)
        # 解析excel文件
        self.parse_excel(filename)
        # 填充stationListWidget
        self.stationListWidget.clear()
        for station in self.stations:
            self.stationListWidget.addItem(str(station['network'] + '/' + station['name']))

    def crawl_btn_clicked(self):
        self.crawl()

        self.hasData=True
        self.currStationIndex=0
        self.currStation = self.stations[self.currStationIndex]
        self.selectedStationChanged()


    def keyPressEvent(self, *args, **kwargs):

        key = args[0].key()
        if key == Qt.Key_A and QApplication.keyboardModifiers() == Qt.ControlModifier:
            self.prior_btn_clicked()
        elif key == Qt.Key_D and QApplication.keyboardModifiers() == Qt.ControlModifier:
            self.next_btn_clicked()

    # 解析excel
    def parse_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet_1 = wb.worksheets[0]  # 根据sheet页的排序选取sheet

        for i in range(2, sheet_1.max_row+1):
            station_name = sheet_1.cell(i, 1).value
            network = sheet_1.cell(i, 2).value
            start_time = sheet_1.cell(i, 3).value
            end_time = sheet_1.cell(i, 4).value
            if not station_name or not network:
                continue
            self.stations.append({
                'name': station_name,
                'network': network,
                'start_time': start_time,
                'end_time': end_time,
                'data': None,
                'select_item': []
            })

        self.crawl_btn.setEnabled(True)
        self.currStationIndex=0
        self.currStation = self.stations[self.currStationIndex]


    def crawl(self):
        if not self.stations:
            QMessageBox.warning(self, "警告", "还没有台站数据",QMessageBox.Ok)
            return

        self.progressDialog = QProgressDialog("爬取进行中，请耐心等待...", "取消", 0, 100, self)
        self.progressDialog.setWindowTitle("爬取进度")
        self.progressDialog.setMinimumSize(240,60)
        #self.progress.setContent("进度", "数据爬取中，请耐心等待")

        self.progressDialog.open()
        QApplication.processEvents()

        for index, station in enumerate(self.stations):
            time.sleep(5)
            station['data'] = crawl(station['name'], station['network'], station['start_time'], station['end_time'], proxies)
            progress_value = (index + 1) / len(self.stations) * 100
            self.progressDialog.setValue(int(progress_value))

            QApplication.processEvents()

        QMessageBox.information(self, "提示", "台站数据爬取完成！", QMessageBox.Ok)
        self.progressDialog.hide()

        return


    def next_btn_clicked(self):
        if not self.hasData:
            QMessageBox.warning(self, "警告", "没有台站工作时间数据，请先爬取",QMessageBox.Ok)
            return

        if self.currStationIndex == len(self.stations)-1 :
            QMessageBox.warning(self,"警告", "已经是最后一个台站了", QMessageBox.Ok)
            return

        self.currStationIndex = self.currStationIndex + 1
        self.currStation = self.stations[self.currStationIndex]

        self.selectedStationChanged()


    def prior_btn_clicked(self):
        if not self.hasData:
            QMessageBox.warning(self, "警告", "没有台站工作时间数据，请先爬取",QMessageBox.Ok)
            return
        if self.currStationIndex == 0:
            QMessageBox.warning(self, "警告", "已经是第一个台站了", QMessageBox.Ok)
            return
        self.currStationIndex=self.currStationIndex-1
        self.currStation = self.stations[self.currStationIndex]
        self.selectedStationChanged()

    def exportSelectedChannels(self, filename):

        with open(filename, 'w', newline='') as csvfile:
            w = csv.writer(csvfile)
            header = ('network', 'name', 'start_time', 'end_time', 'latitude', 'longitude', 'elevation', 'components', 'location_code')
            w.writerow(header)

            row_list = []
            # 相同设备下同赫兹分量合并
            for station in self.stations:
                merge_dict = {}
                for key in station['selectedChannels']:

                    block_index, location, device_name, hertz, c = key.split('||')
                    key = '||'.join([block_index, location, device_name, hertz])
                    if key not in merge_dict.keys():
                        merge_dict[key] = []
                    merge_dict[key].append(c)
                for key, cha_list in merge_dict.items():
                    block_index, location, device_name, hertz = key.split('||')
                    description = station['data'][int(block_index)]['description']
                    row = (station['network'], station['name'], station['start_time'].split('T')[0], station['end_time'].split('T')[0], str(description['latitude']), str(description['longitude']),
                           str(description['elevation']), ' '.join(cha_list), str(location))
                    row_list.append(row)

            for row in row_list:
                w.writerow(row)


    def selectedStationChanged(self):

        self.stationListWidget.setCurrentRow(self.currStationIndex)
        self.desc_label.setText('network: ' + self.currStation['network'] + '\n' + 'station: ' +
                                self.currStation['name'])

        self.treeWidget.itemChanged.disconnect(self.channelSelectionChanged)
        ##add some keys to stations if missing
        if not "allChannels" in self.currStation:
            self.currStation["allChannels" ]={}
        if not "selectedChannels" in self.currStation:
            self.currStation["selectedChannels" ]={}


        self.treeWidget.clear()
        data = self.currStation['data']
        self.component_list.clear()
        # 设置列数
        self.treeWidget.setColumnCount(1)
        # 设置树形控件头部的标题
        self.treeWidget.setHeaderLabels(['节点'])
        for index, content in enumerate(data):
            block = QTreeWidgetItem(self.treeWidget)
            block_text = 'TimePeriod_' + str(index) + '     ' + content['description']['start'] + '   ' + \
                         content['description']['end']
            block.setText(0, block_text)
            #block.setToolTip(0, )
            # locations节点
            for location, device_list in content['locations'].items():
                loc = QTreeWidgetItem(block)
                loc_text = 'location:  ' + location
                loc.setText(0, loc_text)
                # 设备节点
                for device_name, channel_list in device_list.items():
                    device = QTreeWidgetItem(loc)
                    device_text = 'Device:  ' + device_name
                    device.setText(0, device_text)

                    # 频率节点
                    for frequence, channel in channel_list.items():
                        if len(content['locations']) == 1 and len(device_list) == 1 and len(channel_list) == 1:
                            auto_check = True
                        else:
                            auto_check = False
                        hz = QTreeWidgetItem(device)
                        hz.setText(0, frequence)
                        hz.setFlags(hz.flags() | Qt.ItemIsUserCheckable|Qt.ItemIsAutoTristate)
                        #hz.setCheckState(0, Qt.Unchecked)
                        # 分量节点
                        for c in channel.keys():
                            ch = QTreeWidgetItem(hz)
                            ch.setText(0, c)
                            ##using objectId to get all channels
                            tt="objectId:"+str(id(channel[c]))
                            ch.setText(1, tt)
                            self.currStation["allChannels"][tt] = channel[c]

                            tt=content["description"]["start"]+"-"+content["description"]["end"]
                            channel[c]["path"]='/'.join([tt, location, device_name, frequence,c])

                            if not "selected" in channel[c]:  ##add selected key
                                channel[c]["selected"]=Qt.Unchecked
                            ch.setCheckState(0, channel[c]["selected"])

        self.treeWidget.expandAll()

        self.treeWidget.itemChanged.connect(self.channelSelectionChanged)


    def save_btn_clicked(self):
        if not self.hasData:
            QMessageBox.warning(self, "警告", "没有台站工作时间数据，请先爬取",QMessageBox.Ok)
            return
        fileName_choose, filetype = QtWidgets.QFileDialog.getSaveFileName(self, "保存台站信息", os.path.join(os.getcwd(), 'stations.json'),  "json Files (*.json;)")
        if fileName_choose == "":
            return
        self.save_stations(fileName_choose)

    def save_stations(self,fname):
        with open(fname, 'w', encoding='utf-8') as f:
            f.write(json.dumps(self.stations))
        return
    def resetStationListWidget(self):
        self.stationListWidget.clear()
        for station in self.stations:
            if "selectedChannels" in station:
                n=len(station["selectedChannels"])
            else:
                n=0
            tt=station['network'] + '/' + station['name']+" ( "+str(n)+"  channels selected)"
            self.stationListWidget.addItem(tt)



    def load_btn_clicked(self):
        '''
        加载已经下好的数据
        :return:
        '''

        fileName_choose, filetype = QtWidgets.QFileDialog.getOpenFileName(self, "选取文件", os.getcwd(), "Json File (*.json)")
        if fileName_choose == "":
            return

        # 重置软件状态
        self.reset_data()
        with open(fileName_choose, 'r') as f:
            self.stations = json.loads(f.read())
        self.staion_info_path_edit.setText(fileName_choose)
        # 填充stationListWidget
        self.resetStationListWidget()

        self.hasData=True
        self.currStationIndex=0
        self.currStation=self.stations[self.currStationIndex]
        self.selectedStationChanged()


    def station_item_clicked(self, item):
        if not self.hasData:
            QMessageBox.warning(self, "警告", "没有台站工作时间数据，请先爬取",QMessageBox.Ok)
            return
        curr_index = self.stationListWidget.currentRow()
        if curr_index == self.currStationIndex:
            return

        self.currStationIndex = curr_index
        self.currStation = self.stations[self.currStationIndex]
        self.selectedStationChanged()

    def reset_data(self):
        self.stations=[]
        self.currStationIndex = -1
        self.currStation = {}

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    ui = UiMain()
    ui.show()
    sys.exit(app.exec_())
